from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from data import TABLE_TITLE, COLUMNS, TABLE_DATA, FOOTER_NOTE


class TableGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Нормы времени"
        
        # Настройка стилей
        self.header_font = Font(bold=True, size=12, name='Times New Roman')
        self.title_font = Font(bold=True, size=14, name='Times New Roman')
        self.cell_font = Font(size=11, name='Times New Roman')
        
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        
    def create_table(self):
        """Создание таблицы"""
        # Заголовок таблицы
        self.ws.merge_cells('A1:D1')
        title_cell = self.ws['A1']
        title_cell.value = TABLE_TITLE
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment
        
        # Ширина столбцов
        column_widths = [60, 20, 25, 15]
        for i, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(i)
            self.ws.column_dimensions[col_letter].width = width
        
        # Заголовки столбцов
        for col_idx, column_name in enumerate(COLUMNS, 1):
            cell = self.ws.cell(row=2, column=col_idx, value=column_name)
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
            cell.fill = self.header_fill
        
        # Данные таблицы
        for row_idx, row_data in enumerate(TABLE_DATA, 3):
            # Наименование работ
            name_cell = self.ws.cell(row=row_idx, column=1, value=row_data["name"])
            name_cell.font = self.cell_font
            name_cell.alignment = self.left_alignment
            name_cell.border = self.thin_border
            
            # Норма времени
            time_cell = self.ws.cell(row=row_idx, column=2, value=row_data["time_norm"])
            time_cell.font = self.cell_font
            time_cell.alignment = self.center_alignment
            time_cell.border = self.thin_border
            
            # Расценка
            price_cell = self.ws.cell(row=row_idx, column=3, value=row_data["price"])
            price_cell.font = self.cell_font
            price_cell.alignment = self.center_alignment
            price_cell.border = self.thin_border
            
            # Количество
            qty_cell = self.ws.cell(row=row_idx, column=4, value=row_data["quantity"])
            qty_cell.font = self.cell_font
            qty_cell.alignment = self.center_alignment
            qty_cell.border = self.thin_border
        
        # Высота строк
        for row in range(2, len(TABLE_DATA) + 3):
            self.ws.row_dimensions[row].height = 25
        
        # Примечание внизу
        footer_row = len(TABLE_DATA) + 4
        self.ws.merge_cells(f'A{footer_row}:D{footer_row}')
        footer_cell = self.ws[f'A{footer_row}']
        footer_cell.value = FOOTER_NOTE
        footer_cell.font = Font(italic=True, size=11, name='Times New Roman')
        footer_cell.alignment = self.left_alignment
        
        # Установка ориентации страницы альбомной
        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_A4
        
        # Поля страницы
        self.ws.page_margins.left = 0.5
        self.ws.page_margins.right = 0.5
        self.ws.page_margins.top = 0.7
        self.ws.page_margins.bottom = 0.7
        
    def save(self, filename="shop188_norms_2025.xlsx"):
        """Сохранение файла"""
        self.wb.save(filename)
        print(f"✅ Таблица успешно сохранена в файл: {filename}")
