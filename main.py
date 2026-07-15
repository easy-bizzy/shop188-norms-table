from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def create_table():
    """Создание таблицы с формулами для AppSheet"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Нормы цеха 188"
    
    # Заголовок таблицы
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Перечень норм времени и расценок на работы цеха 188 на 2025 год"
    title_cell.font = Font(bold=True, size=14, name='Arial', color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Заголовки столбцов
    headers = ["ID", "Наименование работ", "Норма времени", "Расценка", "Ввод", "Количество"]
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12, name='Arial')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.fill = header_fill
    
    # Данные таблицы
    data = [
        [1, "10-300 Перемещение, укладка тары эл штабелером", 0.28],
        [2, "50-300 Перемещение по скл кран-балкой", 0.34],
        [3, "50-450 Погрузка, выгрузка кран-балкой", 0.34],
        [4, "10-300 Погрузка, выгрузка пленки вручную", 0.56],
        [5, "50-300 Перемещение по скл рохлей", 0.28],
        [6, "50-500 Подъем, перемещение, укладка ПКИ эл штабелером", 0.28],
        [7, "10-300 Подъем, перемещение, укладка ПКИ тележкой", 0.28],
        [8, "50-300 Выгрузка ПКИ из машины эл. штабелером", 0.36],
        [9, "10-50 Погрузка, выгрузка материалов вручную", 0.58],
        [10, "10-50 Погрузка, выгрузка материалов эл штабелером", 0.3],
        [11, "10-50 Погрузка, выгрузка материалов тележкой", 0.36],
        [12, "50-250 Перемещение материалов тележкой", 0.28],
        [13, "50-250 Перемещение материалов эл штабелером", 0.28],
        [14, "10-50 Перемещение материалов вручную", 0.28],
        [15, "10-200 Выгрузка-погрузка ТМЦ вручную", 0.51],
        [16, "10-100 Укладка ТМЦ вручную", 0.42],
        [17, "10-50 Погрузка выгрузка бочек", 0.53],
        [18, "30-300 Выдача ТМЦ в цех рохлей", 0.51],
        [19, "250-500 Выдача ТМЦ в цех с помощью электроштабелера", 0.36],
        [20, "100-500 Выгрузка-погрузка препрега с помощью рохли", 0.30],
        [21, "100-500 Перемещение и укладка препрега с помощью рохли", 0.28],
        [22, "100-300 Погрузка-выгрузка ящиков БПЛА с помощью крана", 0.26],
        [23, "10-300 Укладка, сортировка и перемещение ящиков БПЛА вручную", 0.25],
    ]
    
    # Формула для Google Sheets: суммирует числа разделенные "+"
    # Пример: если D2 = "8+6+4+2+1", то F2 = 21
    formula_template = '=IF(D{row}="";"";ARRAYFORMULA(SUM(VALUE(SPLIT(D{row};"+")))))'
    
    # Заполнение данных
    for row_idx, row_data in enumerate(data, 3):
        # ID
        cell_id = ws.cell(row=row_idx, column=1, value=row_data[0])
        cell_id.font = Font(size=11, name='Arial')
        cell_id.alignment = Alignment(horizontal='center')
        
        # Наименование работ
        cell_name = ws.cell(row=row_idx, column=2, value=row_data[1])
        cell_name.font = Font(size=11, name='Arial')
        cell_name.alignment = Alignment(horizontal='left', wrap_text=True)
        
        # Норма времени
        cell_time = ws.cell(row=row_idx, column=3, value=row_data[2])
        cell_time.font = Font(size=11, name='Arial')
        cell_time.alignment = Alignment(horizontal='center')
        
        # Расценка (пустая, будет заполняться через AppSheet)
        cell_price = ws.cell(row=row_idx, column=4, value="")
        cell_price.font = Font(size=11, name='Arial')
        cell_price.alignment = Alignment(horizontal='center')
        cell_price.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Желтый
        
        # Ввод (пустая, для ввода чисел)
        cell_input = ws.cell(row=row_idx, column=5, value="")
        cell_input.font = Font(size=14, name='Arial', bold=True)
        cell_input.alignment = Alignment(horizontal='center')
        cell_input.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Зеленый
        
        # Количество (формула для автоматического подсчета)
        formula = formula_template.format(row=row_idx)
        cell_quantity = ws.cell(row=row_idx, column=6, value=formula)
        cell_quantity.font = Font(size=12, name='Arial', bold=True, color="006100")
        cell_quantity.alignment = Alignment(horizontal='center')
        cell_quantity.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Светло-зеленый
        
        # Границы для всех ячеек
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Ширина столбцов
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    # Высота строк
    for row in range(2, len(data) + 3):
        ws.row_dimensions[row].height = 30
    
    # Примечание внизу
    footer_row = len(data) + 4
    ws.merge_cells(f'A{footer_row}:F{footer_row}')
    footer_cell = ws[f'A{footer_row}']
    footer_cell.value = "Плёнка, стружка, картон (машиной)"
    footer_cell.font = Font(italic=True, size=11, name='Arial')
    footer_cell.alignment = Alignment(horizontal='left')
    
    # Сохранение файла
    filename = 'shop188_norms_2025.xlsx'
    wb.save(filename)
    print(f"✅ Таблица успешно создана: {filename}")
    print(f"📊 Строк с работами: {len(data)}")
    print(f" Формула добавлена в столбец 'Количество'")
    print(f"🎨 Оформление:")
    print(f"   - Желтый: Расценка (результат)")
    print(f"   - Зеленый: Ввод (поле для ввода)")
    print(f"   - Светло-зеленый: Количество (автосумма)")
    
    return filename


if __name__ == "__main__":
    create_table()
